import json
import re
from typing import Dict, List, Tuple, Any

from openpyxl import load_workbook

from .plane_texts import LABEL_MAPPING_EN, LABEL_MAPPING_ZH

# TASK_PREFIX = "<TASK_ULTRASOUND_DIAGNOSIS> <LEN_SHORT>"
DEFAULT_AUG_ID = 0
DEFAULT_AUG_STRATEGY = "shuffle_pos_tokens_and_planes"

# Strict match examples:
# 1. 四腔心切面（横）
# 1.四腔心切面（横）
PLANE_LINE_RE = re.compile(r"^\s*(\d+)\s*\.\s*(.+?)\s*$")
IMAGE_LINE_RE = re.compile(r"^\s*<image>\s*$")
TOKEN_WRAPPER_RE = re.compile(r"^\s*<\|(.*)\|>\s*$")
TOKEN_RE = re.compile(r"<\|[^<>|]+\|>")


def strip_token_wrapper(s: str) -> str:
    """Remove the outer <| ... |> wrapper."""
    if s is None:
        return ""
    s = str(s).strip()
    m = TOKEN_WRAPPER_RE.match(s)
    if m:
        return m.group(1).strip()
    return s


def normalize_text(s: str) -> str:
    """
    Normalize text for matching:
    1. Keep Chinese parentheses if present
    2. Remove the outer <|...|>
    3. Collapse whitespace
    """
    if s is None:
        return ""
    s = str(s).strip()
    s = strip_token_wrapper(s)
    s = s.replace("\u3000", " ")
    s = re.sub(r"\s+", " ", s).strip()
    return s


def bracket_variants(s: str) -> List[str]:
    """
    Generate bracket variants:
    - original
    - Chinese parentheses -> English parentheses
    - English parentheses -> Chinese parentheses
    """
    s = normalize_text(s)
    if not s:
        return []
    variants = {s}
    variants.add(s.replace("（", "(").replace("）", ")"))
    variants.add(s.replace("(", "（").replace(")", "）"))
    return list(variants)


def separator_variants(s: str) -> List[str]:
    """
    Generate common separator variants for English plane names.
    This helps historical prompts such as 'Tibia/fibula' match
    Excel entries like 'Tibia-fibula' or 'Tibia–fibula'.
    """
    s = normalize_text(s)
    if not s:
        return []

    separators = ["/", "-", "–", "—"]
    variants = {s}

    for old_sep in separators:
        for new_sep in separators:
            variants.add(s.replace(old_sep, new_sep))

    return list(variants)


def all_text_variants(s: str) -> List[str]:
    variants = set(bracket_variants(s))
    variants.update(separator_variants(s))
    return [v for v in variants if v]


def add_mapping_alias(mapping: Dict[str, str], key: str, value: str) -> None:
    key_norm = normalize_text(key)
    if key_norm and key_norm not in mapping:
        mapping[key_norm] = value


def load_plane_mapping(xlsx_path: str) -> Dict[str, str]:
    """
    Load Excel mapping:
      Chinese label -> English_token_real

    Compatible with:
    - Chinese cells like <|四腔心切面（横）|>
    - input text like 四腔心切面（横）
    - both Chinese and English parentheses
    """
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    if "中文" not in headers or "English_token_real" not in headers:
        raise ValueError(
            f"Excel is missing required columns. Expected '中文' and 'English_token_real', got: {headers}"
        )

    idx_cn = headers.index("中文")
    idx_en_label = headers.index("English") if "English" in headers else None
    idx_en_token = headers.index("English_token_real")

    mapping: Dict[str, str] = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        cn = row[idx_cn]
        en = row[idx_en_token]
        if cn is None or en is None:
            continue

        cn_raw = str(cn).strip()
        en_token = str(en).strip()
        if not cn_raw or not en_token:
            continue

        # Raw value
        add_mapping_alias(mapping, cn_raw, en_token)

        # Value after removing <|...|>
        cn_core = normalize_text(cn_raw)
        add_mapping_alias(mapping, cn_core, en_token)

        # Bracket variants
        for alias in all_text_variants(cn_core):
            add_mapping_alias(mapping, alias, en_token)

        # Excel English aliases if present
        if idx_en_label is not None:
            en_label = row[idx_en_label]
            if en_label is not None:
                en_label_raw = str(en_label).strip()
                if en_label_raw:
                    add_mapping_alias(mapping, en_label_raw, en_token)
                    for alias in all_text_variants(en_label_raw):
                        add_mapping_alias(mapping, alias, en_token)

    # Historical aliases from build_report_dataset.py.
    # We first resolve the Chinese labels through Excel, then attach their
    # old English prompt names to the same token to stay backward-compatible.
    for label_key, zh_label in LABEL_MAPPING_ZH.items():
        token = mapping.get(normalize_text(zh_label))
        en_label = LABEL_MAPPING_EN.get(label_key)
        if not token or not en_label:
            continue
        add_mapping_alias(mapping, en_label, token)
        for alias in all_text_variants(en_label):
            add_mapping_alias(mapping, alias, token)

    if not mapping:
        raise ValueError("No valid mapping entries were loaded from Excel.")

    return mapping


def parse_human_content(human_content: str) -> Tuple[List[Tuple[int, str]], str]:
    """
    Parse human content. Expected format:
      1. Four-chamber transverse view
      <image>
      2. Gallbladder view
      <image>
      ...
      Please review...
    Returns:
      planes: [(index, Chinese plane name), ...]
      instruction_text: trailing instruction text
    """
    lines = human_content.splitlines()
    i = 0
    planes: List[Tuple[int, str]] = []

    while i < len(lines):
        line = lines[i].strip()
        m = PLANE_LINE_RE.match(line)
        if not m:
            break

        idx = int(m.group(1))
        plane_cn = normalize_text(m.group(2))

        if i + 1 >= len(lines) or not IMAGE_LINE_RE.match(lines[i + 1].strip()):
            raise ValueError(
                f"Plane {idx} '{plane_cn}' is not immediately followed by <image>."
            )

        planes.append((idx, plane_cn))
        i += 2

    instruction_text = "\n".join(lines[i:]).strip()

    if not planes:
        raise ValueError("No 'index + plane + <image>' structure was parsed.")
    if not instruction_text:
        raise ValueError("Trailing instruction text was not parsed.")

    # Optional: validate consecutive indices
    expected = list(range(1, len(planes) + 1))
    got = [x[0] for x in planes]
    if expected != got:
        raise ValueError(f"Plane indices are not consecutive, expected {expected}, got {got}")

    return planes, instruction_text


def build_target_user_content(
    planes: List[Tuple[int, str]],
    instruction_text: str,
    plane_mapping: Dict[str, str],
) -> str:
    parts = [TASK_PREFIX]

    for idx, plane_cn in planes:
        plane_cn_norm = normalize_text(plane_cn)
        if plane_cn_norm not in plane_mapping:
            raise KeyError(
                f"Plane name was not found in Excel mapping: raw='{plane_cn}', normalized='{plane_cn_norm}'"
            )

        english_token = plane_mapping[plane_cn_norm]
        parts.append(f"{idx}. {english_token}")
        # parts.append(english_token)
        parts.append("<image>")

    parts.append(instruction_text.strip())
    return "\n".join(parts)


def count_tokens(text: str) -> int:
    return len(TOKEN_RE.findall(text or ""))


def get_human_and_gpt(conversations: List[Dict[str, Any]]) -> Tuple[str, str]:
    human_msg = None
    gpt_msg = None

    for conv in conversations:
        if conv.get("from") == "human" and human_msg is None:
            human_msg = conv.get("content", "")
        elif conv.get("from") == "gpt" and gpt_msg is None:
            gpt_msg = conv.get("content", "")

    if human_msg is None:
        raise ValueError("Conversation with from='human' was not found.")
    if gpt_msg is None:
        raise ValueError("Conversation with from='gpt' was not found.")

    return human_msg, gpt_msg


def convert_one_sample(sample: Dict[str, Any], plane_mapping: Dict[str, str]) -> Dict[str, Any]:
    if "image" not in sample:
        raise ValueError("Sample is missing the 'image' field.")
    if "conversations" not in sample:
        raise ValueError("Sample is missing the 'conversations' field.")

    images = sample["image"]
    conversations = sample["conversations"]

    if not isinstance(images, list) or not images:
        raise ValueError("'image' must be a non-empty list.")
    if not isinstance(conversations, list) or not conversations:
        raise ValueError("'conversations' must be a non-empty list.")

    human_msg, gpt_msg = get_human_and_gpt(conversations)
    planes, instruction_text = parse_human_content(human_msg)

    if len(planes) != len(images):
        raise ValueError(
            f"Plane count ({len(planes)}) does not match image count ({len(images)})."
        )

    user_content = build_target_user_content(
        planes=planes,
        instruction_text=instruction_text,
        plane_mapping=plane_mapping,
    )

    assistant_content = str(gpt_msg).strip()

    out = {
        "messages": [
            {
                "role": "user",
                "content": user_content,
            },
            {
                "role": "assistant",
                "content": assistant_content,
            },
        ],
        "images": images,
        "assistant_tokens": assistant_content,
        "user_tokens": "",
        "image_tokens": "",
        "aug_id": DEFAULT_AUG_ID,
        "aug_strategy": DEFAULT_AUG_STRATEGY,
        "symptom_count": count_tokens(assistant_content),
        "plane_count": len(planes),
    }

    # if "id" in sample:
    #     out["id"] = sample["id"]

    return out


def convert_dataset(
    input_json_path: str,
    excel_path: str,
    output_json_path: str,
) -> None:
    plane_mapping = load_plane_mapping(excel_path)

    with open(input_json_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    if not isinstance(data, list):
        raise ValueError("Top-level input JSON must be a list.")

    converted = []
    for i, sample in enumerate(data):
        try:
            converted.append(convert_one_sample(sample, plane_mapping))
        except Exception as e:
            sample_id = sample.get("id", f"index={i}")
            raise RuntimeError(f"Conversion failed for sample {sample_id}: {e}") from e

    with open(output_json_path, "w", encoding="utf-8") as f:
        for obj in converted:
            f.write(json.dumps(obj, ensure_ascii=False) + "\n")
        # json.dump(converted, f, ensure_ascii=False, indent=2)

    print(f"Conversion complete. Total records: {len(converted)}, output: {output_json_path}")


def debug_lookup(term: str, excel_path: str) -> None:
    """Debug whether a Chinese term can be mapped to a token from Excel."""
    mapping = load_plane_mapping(excel_path)
    raw = term
    norm = normalize_text(term)
    print("raw        :", raw)
    print("normalized :", norm)
    print("matched    :", mapping.get(norm))

def main(args):
    # import argparse
    #
    # parser = argparse.ArgumentParser(description="将原始超声脚本输出转换为大模型推理格式")
    # parser.add_argument("--input", default="ultrasound_reports_using_excel.json", help="原始 JSON 路径")
    # parser.add_argument("--excel", default="english_translation_with_tokens.xlsx", help="english_translation_with_tokens.xlsx 路径")
    # parser.add_argument("--output", default="ultrasound_reports_convert.json", help="输出 JSON 路径")
    # parser.add_argument("--task", default="TASK_ULTRASOUND_DIAGNOSIS", help="输出 JSON 路径")
    # parser.add_argument(
    #     "--debug_term",
    #     default=None,
    #     help="可选：调试某个中文术语在 Excel 中是否能匹配到 token，例如 '四腔心切面（横）'",
    # )
    # args = parser.parse_args()

    args.input = args.report["out_json"]
    args.excel = args.convert["en_excel"]
    args.output = args.convert["out_json"]
    args.debug_term = None

    # Determine visit type: prefer patient_info type if available
    visit_type = "Screening" if args.visit_type_is_screening else "Clinical"
    patient_info_path = getattr(args, "patient_info_path", None)
    original_case_name = getattr(args, "original_case_name", "")
    if patient_info_path and original_case_name:
        try:
            import pandas as _pd
            _df = _pd.read_excel(patient_info_path)
            _name_col = _df.columns[0]
            for _, _row in _df.iterrows():
                if str(_row[_name_col]).strip() == original_case_name:
                    _type_val = str(_row.get("type", "")).strip()
                    if _type_val.startswith("<") and _type_val.endswith(">"):
                        visit_type = _type_val[1:-1]
                    break
        except Exception:
            pass

    global TASK_PREFIX
    TASK_PREFIX = f"<{visit_type}> <TASK_ULTRASOUND_REPORT> <LEN_LONG> <{args.center}>" if args.infer_task_is_report else f"<{visit_type}> <TASK_ULTRASOUND_DIAGNOSIS> <LEN_SHORT>"
    if args.debug_term:
        debug_lookup(args.debug_term, args.excel)
    else:
        convert_dataset(args.input, args.excel, args.output)


if __name__ == "__main__":
    pass
